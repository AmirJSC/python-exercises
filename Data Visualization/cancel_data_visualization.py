#! Python3
#  CancelDataVisualization.py - Imports a google sheet and visualizes the data using pygal


'''
What the program does
* Downloads the google sheet as an xlsx file
* Imports the data from the xlsx file into a dictionary
* Use Pygal to visualize the data in Chrome
* Set as a task in windows scheduler every 1st day
of the month and automate the following tasks above.
'''

import logging
import time, datetime
import ezsheets
import openpyxl
import pygal
import os, sys

#Google sheets and drive API have already been activated.

#To enable logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
sheet_id = '1rSmzh7Ld5FiMFv3DoWJ64G_mLh3yAuIdk5yMyRXpWUU'


logging.disable(logging.DEBUG)

#Make a directory in Pythoncode. This will be the cwd.
cwd = 'C:\\Pythoncode\\Cancel Data'
os.makedirs(cwd,exist_ok=True)
os.chdir(cwd)
logging.debug('The current working directory is ' + os.getcwd())


#Check what month it is using the Time module
def get_previous_month():
    current_date = datetime.date.today()
    first_day = current_date.replace(day=1)
    last_month = first_day - datetime.timedelta(days=1)
    last_month = last_month.strftime('%B')
    return last_month


#See if you can access the sheet you're looking for.
try:
    spreadsheet = ezsheets.Spreadsheet(sheet_id)
except:
    print('An exception happened.')
    sys.exit()


sheet_title = list(spreadsheet.sheetTitles)
logging.debug(sheet_title)
last_month = get_previous_month()
logging.debug('The previous month is ' +last_month)
data_file = last_month + '.xlsx'
if last_month not in sheet_title:
    print('There is not sheet named ' +last_month)
    sys.exit()
elif os.path.isfile(data_file) is True:
    print('The data already exists in the ' +cwd +' directory.')
    sys.exit()
else:
    spreadsheet.downloadAsExcel(data_file)

#Delete the other sheets in the excel file
workbook = openpyxl.load_workbook(data_file)
for sheets in [x for x in workbook.sheetnames if x != last_month]:
    del workbook[sheets]
workbook.save(data_file)

#Make a dictionary using the get 'function'
sheet = workbook[last_month]
cancel_count = dict()
for i in range(2,sheet.max_row+1):
    reason = sheet.cell(row=i,column=5).value
    cancel_count[reason] = cancel_count.get(reason,0)+1

logging.debug(cancel_count)

#From the dictionary, plot the data as a bar graph

sorted_cancel_count = sorted(cancel_count.items(), key = lambda kv:kv[1], reverse = True)
sorted_cancel_count = dict(sorted_cancel_count)
print(sorted_cancel_count)
my_config = pygal.Config()
my_config.title_font_size = 24
my_config.show_legend = True
my_config.title_font_size = 24
my_config.label_font_size = 14 
my_config.major_label_font_size = 18
my_config.show_y_guides = True
my_config.width = 800

chart = pygal.Bar(my_config)
chart.title = 'Cancel Data for the month of August'
for label, data_points in sorted_cancel_count.items():
    print(label, data_points)
    chart.add(label, data_points)

chart.render_to_file(last_month+'_cancel_data.svg')










